#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar archivo Excel con listado de recursos
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE_DIR = Path(__file__).parent

def generate_excel():
    """Genera archivo Excel con recursos"""
    
    # Cargar datos
    with open(BASE_DIR / 'recursos_data.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ==========================================
    # HOJA 1: RESUMEN
    # ==========================================
    ws_summary = wb.create_sheet('Resumen', 0)
    
    # Estilos
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="667EEA")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = 'REPORTE GENERAL DE RECURSOS EDUCATIVOS'
    ws_summary['A1'].font = title_font
    ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_summary.row_dimensions[1].height = 25
    
    # Fecha
    ws_summary.merge_cells('A2:D2')
    ws_summary['A2'] = f'Fecha de generación: {datetime.now().strftime("%d de %B de %Y")}'
    ws_summary['A2'].alignment = Alignment(horizontal='center')
    ws_summary.row_dimensions[2].height = 20
    
    ws_summary.row_dimensions[3].height = 5
    
    # Estadísticas generales
    row = 4
    
    # Encabezados
    headers = ['Métrica', 'Valor']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws_summary.row_dimensions[row].height = 20
    row += 1
    
    # Datos
    total_recursos = sum(len(v) for v in data['recursos'].values())
    total_carpetas = len(data['recursos'])
    promedio = round(total_recursos / total_carpetas)
    
    stats = [
        ['Total de Recursos', total_recursos],
        ['Total de Carpetas', total_carpetas],
        ['Promedio por Carpeta', promedio],
        ['Máximo de Recursos', max(len(v) for v in data['recursos'].values())],
        ['Mínimo de Recursos', min(len(v) for v in data['recursos'].values())],
    ]
    
    for stat_row in stats:
        ws_summary.cell(row=row, column=1).value = stat_row[0]
        ws_summary.cell(row=row, column=1).border = border
        ws_summary.cell(row=row, column=2).value = stat_row[1]
        ws_summary.cell(row=row, column=2).font = Font(bold=True)
        ws_summary.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws_summary.cell(row=row, column=2).border = border
        row += 1
    
    # Ancho de columnas
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # ==========================================
    # HOJA 2: DISTRIBUCIÓN POR CARPETA
    # ==========================================
    ws_dist = wb.create_sheet('Distribución', 1)
    
    # Título
    ws_dist.merge_cells('A1:D1')
    ws_dist['A1'] = 'DISTRIBUCIÓN DE RECURSOS POR CARPETA'
    ws_dist['A1'].font = title_font
    ws_dist['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dist.row_dimensions[1].height = 25
    
    ws_dist.row_dimensions[3].height = 5
    
    # Encabezados
    row = 3
    headers = ['#', 'Carpeta', 'Cantidad', 'Porcentaje']
    for col, header in enumerate(headers, 1):
        cell = ws_dist.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws_dist.row_dimensions[row].height = 20
    row += 1
    
    # Datos ordenados
    folders_sorted = sorted(data['recursos'].items(), key=lambda x: len(x[1]), reverse=True)
    
    for idx, (folder, resources) in enumerate(folders_sorted, 1):
        count = len(resources)
        percentage = (count / total_recursos) * 100
        
        ws_dist.cell(row=row, column=1).value = idx
        ws_dist.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws_dist.cell(row=row, column=1).border = border
        
        ws_dist.cell(row=row, column=2).value = folder
        ws_dist.cell(row=row, column=2).border = border
        
        ws_dist.cell(row=row, column=3).value = count
        ws_dist.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws_dist.cell(row=row, column=3).border = border
        ws_dist.cell(row=row, column=3).font = Font(bold=True)
        
        ws_dist.cell(row=row, column=4).value = f'{percentage:.1f}%'
        ws_dist.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws_dist.cell(row=row, column=4).border = border
        
        row += 1
    
    # Ancho de columnas
    ws_dist.column_dimensions['A'].width = 5
    ws_dist.column_dimensions['B'].width = 25
    ws_dist.column_dimensions['C'].width = 15
    ws_dist.column_dimensions['D'].width = 15
    
    # ==========================================
    # HOJA 3: LISTADO COMPLETO DE RECURSOS
    # ==========================================
    ws_list = wb.create_sheet('Recursos', 2)
    
    # Título
    ws_list.merge_cells('A1:D1')
    ws_list['A1'] = 'LISTADO COMPLETO DE RECURSOS'
    ws_list['A1'].font = title_font
    ws_list['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_list.row_dimensions[1].height = 25
    
    ws_list.row_dimensions[3].height = 5
    
    # Encabezados
    row = 3
    headers = ['Carpeta', 'Nombre del Recurso', 'Archivo', 'Ruta']
    for col, header in enumerate(headers, 1):
        cell = ws_list.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
    
    ws_list.row_dimensions[row].height = 20
    row += 1
    
    # Datos
    for folder in sorted(data['recursos'].keys()):
        for resource in sorted(data['recursos'][folder], key=lambda x: x['nombre']):
            ws_list.cell(row=row, column=1).value = folder
            ws_list.cell(row=row, column=1).border = border
            
            ws_list.cell(row=row, column=2).value = resource['nombre'].replace('_', ' ')
            ws_list.cell(row=row, column=2).border = border
            
            ws_list.cell(row=row, column=3).value = resource['archivo']
            ws_list.cell(row=row, column=3).border = border
            
            ws_list.cell(row=row, column=4).value = resource['ruta']
            ws_list.cell(row=row, column=4).border = border
            
            row += 1
    
    # Ancho de columnas
    ws_list.column_dimensions['A'].width = 20
    ws_list.column_dimensions['B'].width = 45
    ws_list.column_dimensions['C'].width = 40
    ws_list.column_dimensions['D'].width = 35
    
    # Guardar archivo
    output_path = BASE_DIR / 'recursos_data.xlsx'
    wb.save(output_path)
    
    print(f"✅ Archivo Excel generado exitosamente: {output_path}")
    print(f"   - Hoja 'Resumen': Estadísticas generales")
    print(f"   - Hoja 'Distribución': Recursos por carpeta")
    print(f"   - Hoja 'Recursos': Listado completo ({total_recursos} recursos)")

if __name__ == '__main__':
    generate_excel()
